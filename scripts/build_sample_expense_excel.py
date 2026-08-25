import sys, os, openpyxl, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

src_path = r'C:\Users\AMD\Desktop\Container_V3_WebApp\sample_data\May-69_รถหัวลากรับ-จ่าย (1).xlsx'
out_path = r'C:\Users\AMD\Desktop\Container_V3_WebApp\sample_data\ตัวอย่างนำเข้า_ค่าใช้จ่ายรถ_พฤษภาคม2569.xlsx'

wb_src = openpyxl.load_workbook(src_path, data_only=True)

def detect_category(desc):
    t = str(desc or '').lower().strip()
    if not t: return 'misc'
    if 'เติมน้ำมัน' in t or 'น้ำมัน' in t or 'ดีเซล' in t: return 'fuel'
    if any(k in t for k in ['ซ่อม', 'ปะยาง', 'เปลี่ยนยาง', 'สลับยาง', 'แอร์', 'ฟิล์ม', 'ถ่ายน้ำมัน', 'กลอน', 'วาล์ว', 'กรอง', 'ไฟ', 'ครัช', 'เบรค', 'จารบี', 'ลมยาง', 'ช่าง', 'อู่', 'พัดลม', 'ลูกยาง', 'ท่อยาง', 'ปั้ม']):
        return 'maintenance'
    if any(k in t for k in ['ผ่านท่า', 'ทางด่วน', 'มอเตอร์ไซด์', 'มอไซด์', 'สะพาน', 'ที่จอด']):
        return 'toll_port'
    if any(k in t for k in ['ผ่อนรถ', 'ผ่อนหาง', 'งวด', 'ค่างวด']):
        return 'installment'
    if any(k in t for k in ['ทะเบียน', 'พรบ', 'พ.ร.บ', 'ประกัน', 'ตรวจสภาพ', 'ภาษี']):
        return 'tax_insurance'
    return 'misc'

CAT_LABEL = {
    'fuel': 'ค่าน้ำมัน',
    'maintenance': 'ซ่อมบำรุง & อะไหล่',
    'toll_port': 'ค่าผ่านทาง / ผ่านท่า',
    'installment': 'ค่างวดรถ & หาง',
    'tax_insurance': 'ภาษี & พ.ร.บ. & ประกัน',
    'misc': 'จิปาถะ / กองกลาง'
}

all_records = []

for sname in wb_src.sheetnames:
    if any(skip in sname for skip in ['ร้านโชห่วย', 'รายรับ-จ่าย จริง', 'ทะเบียน ประกัน', 'template']):
        continue
    ws = wb_src[sname]
    
    # default truck & driver
    default_truck = 'FLEET_SHARED'
    default_driver = 'กองกลาง'
    batch_name = 'พฤษภาคม 2569'
    
    if sname == 'รายจ่ายอื่นๆ':
        default_truck = 'FLEET_SHARED'
        default_driver = 'กองกลาง'
    else:
        import re
        m = re.search(r'(\d+)', sname)
        if m: default_truck = m.group(1)
        r2_val = ws.cell(2, 3).value
        if r2_val: default_driver = str(r2_val).strip()
        else: default_driver = sname

    for r in range(6, ws.max_row + 1):
        c1 = ws.cell(r, 1).value # date
        c2 = ws.cell(r, 2).value # desc
        c3 = ws.cell(r, 3).value # trip count
        c6 = ws.cell(r, 6).value # goods / buy
        c7 = ws.cell(r, 7).value # labor
        c11 = ws.cell(r, 11).value or ws.cell(r, 10).value # doc / remark
        
        if not c2 or not str(c2).strip():
            continue
        desc = str(c2).strip()
        
        # skip payroll & summary
        if any(k in desc for k in ['เบิกค่าเที่ยว', 'จ่ายเงินเดือน', 'ประกันสังคม', 'หักส่วนบุคคล', 'หักเงินยืม', 'จ่ายพนักงาน', 'รวมรายจ่าย', 'กำไร', 'คำนวณค่าตู้', 'สรุปรายได้', 'หักค่าผ่อนกระบะ', 'เงินพิเศษ']):
            continue
            
        g_val = float(c6) if isinstance(c6, (int, float)) else 0.0
        l_val = float(c7) if isinstance(c7, (int, float)) else 0.0
        trip_val = float(c3) if isinstance(c3, (int, float)) else 0.0
        tot_val = g_val + l_val
        if tot_val == 0 and trip_val == 0:
            continue
            
        # Parse date
        date_str = '2026-05-01'
        if isinstance(c1, datetime.datetime) or isinstance(c1, datetime.date):
            y = c1.year
            if y > 2500: y -= 543
            date_str = f'{y:04d}-{c1.month:02d}-{c1.day:02d}'
        elif isinstance(c1, str) and '/' in c1:
            parts = c1.strip().split('/')
            if len(parts) == 3:
                try:
                    y = int(parts[2])
                    if y > 2500: y -= 543
                    date_str = f'{y:04d}-{int(parts[1]):02d}-{int(parts[0]):02d}'
                except: pass
                
        cat = detect_category(desc)
        cost_per_trip = round(tot_val / trip_val, 2) if trip_val > 0 else 0.0
        
        vendor = '-'
        if 'ผ่านท่า' in desc: vendor = 'ผ่านท่า'
        elif 'ช่างเอ' in desc: vendor = 'ช่างเอ'
        elif 'Makro' in desc: vendor = 'Makro'
        elif 'Lotus' in desc: vendor = "Lotus's"
        
        remark = str(c11).strip() if c11 and str(c11).strip() != 'None' else '-'
        
        all_records.append({
            'date': date_str,
            'truck_no': default_truck,
            'driver_name': default_driver,
            'batch_name': batch_name,
            'category': cat,
            'category_label': CAT_LABEL[cat],
            'description': desc,
            'amount_goods': g_val,
            'amount_labor': l_val,
            'amount_total': tot_val,
            'trip_count': trip_val,
            'cost_per_trip': cost_per_trip,
            'vendor_name': vendor,
            'invoice_no': f'INV-6905-{len(all_records)+1:03d}',
            'payment_method': 'บัตรน้ำมัน' if cat == 'fuel' else ('เงินสด' if tot_val < 2000 else 'โอนเงิน'),
            'remark': remark,
            'sheet_origin': sname
        })

print(f'Successfully parsed {len(all_records)} expense records from original Excel!')

# Write to a clean new workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'รายการค่าใช้จ่ายรถ_พฤษภาคม2569'
ws_out.views.sheetView[0].showGridLines = True

headers = [
    'ลำดับ', 'วันที่', 'เบอร์รถ', 'ชื่อคนขับ', 'งวดงาน/เดือน', 
    'รหัสหมวดหมู่', 'หมวดหมู่ค่าใช้จ่าย', 'รายการค่าใช้จ่าย', 
    'ค่าของ/อะไหล่/น้ำมัน (บาท)', 'ค่าแรงช่าง (บาท)', 'ยอดรวมสุทธิ (บาท)', 
    'จน.เที่ยวที่วิ่ง (เที่ยว)', 'ค่าน้ำมันเฉลี่ย/เที่ยว (บาท)', 
    'ร้านค้า/อู่/ปั๊ม', 'เลขที่บิล/ใบเสร็จ', 'วิธีชำระเงิน', 'หมายเหตุ', 'ชีทเดิมอ้างอิง'
]

ws_out.append(headers)

# Style Header
header_font = Font(name='Sarabun', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy Blue
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

ws_out.row_dimensions[1].height = 28
for col_num in range(1, len(headers) + 1):
    c = ws_out.cell(1, col_num)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

# Append Data
for idx, r in enumerate(all_records, 1):
    row_data = [
        idx,
        r['date'],
        'กองกลาง' if r['truck_no'] == 'FLEET_SHARED' else r['truck_no'],
        r['driver_name'],
        r['batch_name'],
        r['category'],
        r['category_label'],
        r['description'],
        r['amount_goods'],
        r['amount_labor'],
        r['amount_total'],
        r['trip_count'] if r['trip_count'] > 0 else '',
        r['cost_per_trip'] if r['cost_per_trip'] > 0 else '',
        r['vendor_name'],
        r['invoice_no'],
        r['payment_method'],
        r['remark'],
        r['sheet_origin']
    ]
    ws_out.append(row_data)
    row_idx = idx + 1
    ws_out.row_dimensions[row_idx].height = 22
    
    # Alternating row fill
    fill_color = 'F8FAFC' if idx % 2 == 0 else 'FFFFFF'
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    for c_idx in range(1, len(headers) + 1):
        cell = ws_out.cell(row_idx, c_idx)
        cell.font = Font(name='Sarabun', size=10)
        cell.fill = row_fill
        cell.border = thin_border
        
        # Alignment & Formatting
        if c_idx in [1, 2, 3, 5, 6]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c_idx in [9, 10, 11, 12, 13]:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if c_idx in [9, 10, 11]:
                cell.number_format = '#,##0.00'
            elif c_idx == 12:
                cell.number_format = '#,##0.0'
            elif c_idx == 13:
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Summary Row at Bottom
total_row_idx = len(all_records) + 2
ws_out.cell(total_row_idx, 1, 'รวมทั้งหมด')
ws_out.cell(total_row_idx, 1).font = Font(name='Sarabun', size=11, bold=True)
ws_out.cell(total_row_idx, 1).alignment = Alignment(horizontal='center', vertical='center')
ws_out.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=8)

for col_letter, num_idx in [('I', 9), ('J', 10), ('K', 11), ('L', 12)]:
    c = ws_out.cell(total_row_idx, num_idx)
    c.value = f'=SUM({col_letter}2:{col_letter}{total_row_idx-1})'
    c.font = Font(name='Sarabun', size=11, bold=True, color='1E40AF')
    c.number_format = '#,##0.00' if num_idx != 12 else '#,##0.0'
    c.alignment = Alignment(horizontal='right', vertical='center')

total_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
for c_idx in range(1, len(headers) + 1):
    cell = ws_out.cell(total_row_idx, c_idx)
    cell.fill = total_fill
    cell.border = Border(
        top=Side(style='thin', color='1E40AF'),
        bottom=Side(style='double', color='1E40AF'),
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1')
    )

# Auto-fit column widths
for col in ws_out.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb_out.save(out_path)
print(f'Successfully saved clean sample file to: {out_path}')
